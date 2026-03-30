from openpyxl import load_workbook
from collections import defaultdict


def normalize_port(port):
    if not port:
        return ""
    return str(port).split(",")[0].strip().upper()


def build_pod_to_rows(sheet, start_row, end_row, pod_col="F"):
    pod_to_rows = defaultdict(list)

    for row in range(start_row, end_row + 1):
        pod_value = sheet[f"{pod_col}{row}"].value
        pod_key = normalize_port(pod_value)
        if pod_key:
            pod_to_rows[pod_key].append(row)

    return dict(pod_to_rows)


def update_whf_rates(sheet, whf_rates):

    # 建立 POD → row mapping（全部區間）
    pod_to_rows = build_pod_to_rows(sheet, 5, 98, pod_col="F")

    # 找出會被更新的 rows
    rows_to_update = set()
    for city in whf_rates.keys():
        rows = pod_to_rows.get(city, [])
        rows_to_update.update(rows)

    # 先清掉 QRS（避免殘值）
    for row in rows_to_update:
        sheet[f"Q{row}"] = None
        sheet[f"R{row}"] = None
        sheet[f"S{row}"] = None

    # 寫入 WHF
    for city, rate_dict in whf_rates.items():

        target_rows = pod_to_rows.get(city, [])

        for row in target_rows:
            sheet[f"Q{row}"] = rate_dict["20"]
            sheet[f"R{row}"] = rate_dict["40"]
            sheet[f"S{row}"] = rate_dict["40HC"]


def update_excel_rates_test(excel_path, rates, whf_rates):
    # 👉 讀「公式」
    wb = load_workbook(excel_path)
    sheet = wb.active

    # 👉 讀「計算後的值」
    wb_value = load_workbook(excel_path, data_only=True)
    sheet_value = wb_value.active

    # 1. 備份舊值（用 value workbook）
    for row in range(5, 98):
        sheet[f"AD{row}"] = sheet_value[f"W{row}"].value
        sheet[f"AE{row}"] = sheet_value[f"X{row}"].value
        sheet[f"AF{row}"] = sheet_value[f"Y{row}"].value

    # 2. 依照目前 Excel 內容動態建立 mapping
    bremerhaven_pod_to_rows = build_pod_to_rows(sheet, 5, 11, pod_col="F")
    hamburg_pod_to_rows = build_pod_to_rows(sheet, 26, 28, pod_col="F")

    # 3. 先只清掉會被更新到的 row，避免殘值
    rows_to_clear = set()
    for rows in bremerhaven_pod_to_rows.values():
        rows_to_clear.update(rows)
    for rows in hamburg_pod_to_rows.values():
        rows_to_clear.update(rows)

    for row in rows_to_clear:
        sheet[f"H{row}"] = None
        sheet[f"I{row}"] = None
        sheet[f"J{row}"] = None

    # 4. 寫 BREMERHAVEN / HAMBURG
    for r in rates:
        pol = r["POL"].strip().upper()
        pod = normalize_port(r["POD"])

        if pol == "BREMERHAVEN":
            target_rows = bremerhaven_pod_to_rows.get(pod, [])
        elif pol == "HAMBURG":
            target_rows = hamburg_pod_to_rows.get(pod, [])
        else:
            continue

        for row in target_rows:
            sheet[f"H{row}"] = r["20"]
            sheet[f"I{row}"] = r["40"]
            sheet[f"J{row}"] = r["40HC"]
    
    # 5. 寫 WHF
    update_whf_rates(sheet, whf_rates)


    wb.save(excel_path)

